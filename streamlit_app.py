import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import zipfile
import io
import openpyxl
from pandas import ExcelWriter
from streamlit_sortables import sort_items
from scipy.stats import norm
from matplotlib.patches import FancyBboxPatch
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook

# Charger le fichier Excel
file_path = 'NORMES_NOV_24.xlsx'
excel_data = pd.ExcelFile(file_path)

# Liste des groupes d'âge (onglets du fichier)
age_groups = excel_data.sheet_names

if "age_selected" not in st.session_state:
    st.session_state["age_selected"] = False

if "scores_entered" not in st.session_state:
    st.session_state["scores_entered"] = False

if "age_data" not in st.session_state:
    st.session_state["age_data"] = pd.DataFrame()

if "missing_norms" not in st.session_state:
    st.session_state["missing_norms"] = []

st.markdown(
    """
    <div style="text-align: center; font-size: 40px; font-weight: bold;">
        Batterie COMPRENDRE
    </div>
    """,
    unsafe_allow_html=True
)

#Âge ET ID
st.header("Étape 1 : Sélectionnez le groupe d'âge")
selected_age_group = st.selectbox("Sélectionnez le groupe d'âge de l'enfant :", age_groups)
child_id = st.text_input("Saisissez l'ID de l'enfant :", value="", placeholder="ID de l'enfant")

if st.button("Passer à l'étape suivante"):
    if not child_id.strip(): 
        st.error("Veuillez saisir un ID valide avant de continuer.")
    else:
        st.session_state["age_selected"] = True
        st.session_state["child_id"] = child_id 
        st.success(f"ID {child_id} et âge {selected_age_group} confirmés.")
        
        
def load_age_data(sheet_name, excel_file):
    try:
        return pd.read_excel(excel_file, sheet_name=sheet_name, engine="openpyxl")
    except Exception as e:
        st.error(f"Erreur lors du chargement des données : {e}")
        return pd.DataFrame()


if st.session_state["age_selected"]:
    st.header("Étape 2 : Entrez les scores")
    age_data = load_age_data(selected_age_group, excel_data)

    if age_data.empty:
        st.error("Impossible de charger les données pour le groupe d'âge sélectionné.")
    else:
        age_data = age_data[["Tâche", "Moyenne", "Ecart-type", "Minimum", 
                             "5e percentile", "10e percentile", "Q1", 
                             "Q2 - mediane", "Q3", "90e percentile", "Maximum"]].dropna()

        # Liste des catégories avec les tâches regroupées par paires
        categories = {
            "Langage": [
                ("Discrimination Phonologique", "Décision Lexicale Auditive"),
                ("Mots Outils", "Stock Lexical"),
                ("Compréhension Syntaxique", "Mots Outils - BOEHM")
            ],
            "Mémoire de Travail Verbale": [
                ("Mémoire de travail verbale endroit empan", "Mémoire de travail verbale endroit brut"),
                ("Mémoire de travail verbale envers empan", "Mémoire de travail verbale envers brut")
            ],
            "Mémoire de Travail Non Verbale": [
                ("Mémoire de travail non verbale endroit empan", "Mémoire de travail non verbale endroit brut"),
                ("Mémoire de travail non verbale envers empan", "Mémoire de travail non verbale envers brut")  
            ],   
            "Mise à jour Verbale": [
                ("Mise à jour verbale empan", "Mise à jour verbale score"),
            ],
            "Mise à jour Non Verbale": [
                ("Mise à jour non verbale empan", "Mise à jour non verbale score"),
            ],
            "INHIB verbale": [
                ("Inhibition verbale congruent score", "Inhibition verbale incongruent score"),
                ("Inhibition verbale congruent temps", "Inhibition verbale incongruent temps")
            ],
            "INHIB non verbale": [
                ("Inhibition non verbale congruent score", "Inhibition non verbale incongruent score"),
                ("Inhibition non verbale congruent temps", "Inhibition non verbale incongruent temps")
            ]
        }

        # Collecte des scores utilisateur et calculs d'interférences
        user_scores = []
        inhibition_scores = {}
        missing_norms = []

        for category, task_pairs in categories.items():
            st.subheader(category)
            for task1, task2 in task_pairs:
                col1, col2 = st.columns(2)

                # Colonne 1 : Saisie pour task1
                with col1:
                    if task1 in age_data["Tâche"].values:
                        score1 = st.text_input(f"{task1} :", value="")
                        if score1.strip(): 
                            try:
                                score1 = float(score1)
                                user_scores.append({"Tâche": task1, "Score Enfant": score1})
                                inhibition_scores[task1] = score1
                            except ValueError:
                                st.error(f"Valeur non valide pour {task1}. Veuillez entrer un nombre.")
                                inhibition_scores[task1] = score1
                    else:
                        st.warning(f"Pas de normes disponibles pour {task1}")
                        missing_norms.append(task1)

                # Colonne 2 : Saisie pour task2
                with col2:
                    if task2 in age_data["Tâche"].values:
                        score2 = st.text_input(f"{task2} :", value="")
                        if score2.strip():  
                            try:
                                score2 = float(score2)
                                user_scores.append({"Tâche": task2, "Score Enfant": score2})
                                inhibition_scores[task2] = score2
                            except ValueError:
                                st.error(f"Valeur non valide pour {task2}. Veuillez entrer un nombre.")
                                inhibition_scores[task2] = score2
                    else:
                        st.warning(f"Pas de normes disponibles pour {task2}")
                        missing_norms.append(task2)

        # Calculs des interférences
        interferences = {
            "Inhibition verbale interférence score": (
                inhibition_scores.get("Inhibition verbale incongruent score", 0) 
                - inhibition_scores.get("Inhibition verbale congruent score", 0)
            ),
            "Inhibition non verbale interférence score": (
                inhibition_scores.get("Inhibition non verbale incongruent score", 0) 
                - inhibition_scores.get("Inhibition non verbale congruent score", 0)
            ),
            "Inhibition verbale interférence temps": (
                inhibition_scores.get("Inhibition verbale incongruent temps", 0) 
                - inhibition_scores.get("Inhibition verbale congruent temps", 0)
            ),
            "Inhibition non verbale interférence temps": (
                inhibition_scores.get("Inhibition non verbale incongruent temps", 0) 
                - inhibition_scores.get("Inhibition non verbale congruent temps", 0)
            )
        }

        st.subheader("Scores d'interférence calculés")
        for key, value in interferences.items():
            st.write(f"**{key}** : {value:.2f}")

       
        filtered_interferences = {
            key: value for key, value in interferences.items() if value != 0
        }

        user_scores.extend(
            [{"Tâche": key, "Score Enfant": value} for key, value in filtered_interferences.items()]
        )

        scores_df = pd.DataFrame(user_scores, columns=["Tâche", "Score Enfant"])

        # Fusionner avec les données originales pour les calculs
        merged_data = pd.merge(age_data, scores_df, on="Tâche", how="left")
        merged_data["Z-Score"] = (merged_data["Score Enfant"] - merged_data["Moyenne"]) / merged_data["Ecart-type"]
        merged_data["Z-Score"] = pd.to_numeric(merged_data["Z-Score"], errors="coerce")
        merged_data = merged_data.dropna(subset=["Z-Score"])

        # Inverser les Z-scores pour les variables de temps d'inhibition
        time_variables = [
            "Inhibition verbale congruent temps",
            "Inhibition verbale incongruent temps",
            "Inhibition non verbale congruent temps",
            "Inhibition non verbale incongruent temps", 
            "Inhibition verbale interférence temps", 
            "Inhibition non verbale interférence temps"
        ]
        
        merged_data["Percentile (%)"] = norm.cdf(merged_data["Z-Score"]) * 100

        filled_data = merged_data[~merged_data["Score Enfant"].isna()]
        filled_data = filled_data.drop_duplicates(subset="Tâche")

        # Bouton pour confirmer les scores
        if st.button("Confirmer les scores et afficher les résultats"):
            st.session_state["scores_entered"] = True
            st.session_state["age_data"] = filled_data
            st.session_state["missing_norms"] = missing_norms

# Étape 3 : Résultats
    categories_mapping = {
        "Langage": [
            "Discrimination Phonologique", "Décision Lexicale Auditive",
            "Mots Outils", "Stock Lexical", "Compréhension Syntaxique", "Mots Outils - BOEHM"
        ],
        "Mémoire de Travail": [
            "Mémoire de travail verbale endroit empan", "Mémoire de travail verbale endroit brut",
            "Mémoire de travail verbale envers empan", "Mémoire de travail verbale envers brut",
            "Mémoire de travail non verbale endroit empan", "Mémoire de travail non verbale endroit brut",
            "Mémoire de travail non verbale envers empan", "Mémoire de travail non verbale envers brut"
        ],
        "Mise à jour": [
            "Mise à jour verbale empan", "Mise à jour verbale score",
            "Mise à jour non verbale empan", "Mise à jour non verbale score"
        ],
        "Inhibition": [
            "Inhibition verbale congruent score", "Inhibition verbale incongruent score",
            "Inhibition verbale congruent temps", "Inhibition verbale incongruent temps",
            "Inhibition verbale interférence score", "Inhibition verbale interférence temps",
            "Inhibition non verbale congruent score", "Inhibition non verbale incongruent score",
            "Inhibition non verbale congruent temps", "Inhibition non verbale incongruent temps",
            "Inhibition non verbale interférence score", "Inhibition non verbale interférence temps"
        ]
    }

    task_name_mapping = {
        "Discrimination Phonologique": "Discrimination\nPhonologique",
        "Décision Lexicale Auditive": "Décision\nLexicale\nAuditive",
        "Mots Outils": "Mots\nOutils",
        "Stock Lexical": "Stock\nLexical",
        "Compréhension Syntaxique": "Compréhension\nSyntaxique",
        "Mots Outils - BOEHM": "BOEHM",
        "Mémoire de travail verbale endroit empan": "Mémoire de travail\nVebrale\nendroit\nempan",
        "Mémoire de travail verbale endroit brut": "Mémoire de travail\nVerbale\nendroit\nbrut",
        "Mémoire de travail verbale envers empan": "Mémoire de travail\nVerbale\nenvers\nempan",
        "Mémoire de travail verbale envers brut": "Mémoire de travail\nVerbale\nenvers\nbrut",
        "Mémoire de travail non verbale endroit empan": "Mémoire de travail\nNon Verbale\nendroit\nempan",
        "Mémoire de travail non verbale endroit brut": "Mémoire de travail\nNon Verbale\nendroit\nbrut",
        "Mémoire de travail non verbale envers empan": "Mémoire de travail\nNon Verbale\nenvers\nempan",
        "Mémoire de travail non verbale envers brut": "Mémoire de travail\nNon Verbale\nenvers\nbrut",
        "Mise à jour verbale empan": "Mise-à-jour\nVerbale\nempan",
        "Mise à jour verbale score": "Mise-à-jour\nVerbale\nbrut",
        "Mise à jour non verbale empan": "Mise-à-jour\nNon Verbale\nempan",
        "Mise à jour non verbale score": "Mise-à-jour\nNon Verbale\nbrut",
        "Inhibition verbale congruent score": "Inhibition\nVerbale\nCongruent\nscore",
        "Inhibition verbale incongruent score": "Inhibition\nVerbale\nIncongruent\nscore",
        "Inhibition verbale congruent temps": "Inhibition\nVerbale\nCongruent\ntemps",
        "Inhibition verbale incongruent temps": "Inhibition\nVerbale\nIncongruent\ntemps",
        "Inhibition verbale interférence score": "Inhibition\nVerbale\nscore",
        "Inhibition verbale interférence temps": "Inhibition\nVerbale\ntemps",
        "Inhibition non verbale congruent score": "Inhibition\nNon Verbale\nCongruent\nscore",
        "Inhibition non verbale incongruent score": "Inhibition\nNon Verbale\nIncongruent\nscore",
        "Inhibition non verbale congruent temps": "Inhibition\nNon Verbale\nCongruent\ntemps",
        "Inhibition non verbale incongruent temps": "Inhibition\nNon Verbale\nIncongruent\ntemps",
        "Inhibition non verbale interférence score": "Inhibition\nNon Verbale\nscore",
        "Inhibition non verbale interférence temps": "Inhibition\nNon Verbale\ntemps"
    }

    # Ajouter la colonne "Catégorie" pour chaque tâche
    def plot_grouped_scores(data, selected_tasks):
        category_colors = {
            "Langage": "#3798da",
            "Mémoire de Travail": "#eca113",
            "Mise à jour": "#e365d6",
            "Inhibition": "#8353da",
            "Autre": "gray"
        }

        # Filtrer les données pour inclure uniquement les tâches sélectionnées
        data = data[data["Tâche"].isin(selected_tasks)]

        # Liste des tâches (abrégées) et leurs Z-scores
        tasks = data["Tâche"].map(task_name_mapping).tolist()
        percentiles = data["Percentile (%)"].tolist()

        positions = np.arange(len(tasks))  

        # Ajouter une colonne pour les positions dans le DataFrame
        data["Position"] = positions

        # Créer la figure
        fig_width = 14
        fig_height = max(10, len(tasks) * 1.5)
        fig, ax = plt.subplots(figsize=(fig_width, fig_height))

        # Tracer les points pour chaque tâche
        point_colors = data["Catégorie"].map(category_colors)
        ax.scatter(percentiles, positions, color=point_colors, s=100, zorder=3)

# Ajouter les scores de l'enfant avec un cadre coloré autour
        for i, (score, category) in enumerate(zip(data["Score Enfant"], data["Catégorie"])):
            color = category_colors.get(category, "gray")  

            # Calculer la hauteur en fonction de l'espacement des points sur l'axe Y
            if len(positions) > 1:  
                spacing = positions[1] - positions[0]  
            else:
                spacing = 1  

            box_height = spacing*0.001  # Ajuster la hauteur proportionnellement à l'espacement
            vertical_offset = box_height / 2  # Centrer le cadre autour du point

            # Ajouter le cadre
            bbox = FancyBboxPatch(
                (105, positions[i] - vertical_offset),  # Coordonnées (x, y) centrées
                width=8,  # Largeur du cadre
                height=box_height,  # Hauteur ajustée dynamiquement
                boxstyle="square,pad=0.1",  # Angles arrondis avec padding
                linewidth=2,  # Épaisseur de la bordure
                edgecolor=color,  # Couleur de la bordure
                facecolor="white",  # Couleur de fond
                zorder=1  # Couche d'affichage
            )
            ax.add_patch(bbox)  # Ajouter le cadre au graphique

            # Ajouter le texte centré dans le cadre
            ax.text(
                x=109,  # Position X centrée dans le cadre
                y=positions[i],  # Position Y alignée verticalement au centre
                s=f"{score:.0f}",  # Le score formaté
                fontsize=14,
                fontweight = "bold", 
                color="black",  # Couleur du texte
                ha="center",  # Alignement horizontal centré
                va="center",  # Alignement vertical centré
                zorder=2  # Couche d'affichage au-dessus du cadre
            )

        # Ajouter des zones colorées pour les catégories
        ax.fill_betweenx(range(-1, len(tasks)+1), 0, 3, color="#d44646", alpha=0.2, zorder=1)  # Zone rouge
        ax.fill_betweenx(range(-1, len(tasks)+1), 3, 15, color="#f5a72f", alpha=0.2, zorder=1)  # Zone orange
        ax.fill_betweenx(range(-1, len(tasks)+1), 15, 85, color="#60cd72", alpha=0.2, zorder=1)  # Zone verte
        ax.fill_betweenx(range(-1, len(tasks)+1), 85, 97, color="#8ddf9b", alpha=0.2, zorder=1)  # Zone vert clair
        ax.fill_betweenx(range(-1, len(tasks)+1), 97, 100, color="#aedeb6", alpha=0.2, zorder=1)  # Zone bleue


        # Ligne de référence Z=0
        ax.axvline(50, color="black", linestyle="--", linewidth=0.8, zorder=2)
        
        ax.set_xlim(0, 120)  # Axe X : percentiles de 0 à 100
        ax.set_ylim(-1, len(tasks))

        # Configurer les ticks et les labels
        ax.set_xticks([0, 3, 15, 50, 85, 97, 100])
        ax.set_xticklabels(["0", "3", "15", "50", "85", "97", "100"], fontsize=12, fontweight="bold", rotation = -35)
        ax.set_yticks(positions)
        ax.set_yticklabels(tasks, fontsize=16, fontweight="bold")
        ax.set_xlabel("Percentiles (%)", fontsize=14)
        ax.xaxis.set_label_coords(0.43, -0.05)
        ax.set_ylabel("")

        fig.suptitle(
            "Résultats Batterie Comprendre",
            fontsize=24,
            fontweight="bold",
            x= 0.53, 
            y=1       
        )

        for idx, category in enumerate(category_colors.keys()):
            # Filtrer les données pour cette catégorie
            category_data = data[data["Catégorie"] == category]
            
            # Obtenir les positions et les percentiles pour les tâches dans la catégorie
            category_positions = category_data["Position"].tolist() if not category_data.empty else []
            category_percentiles = category_data["Percentile (%)"].tolist() if not category_data.empty else []

            # Relier les points avec une ligne si la catégorie n'est pas vide
            if category_positions and category_percentiles:
                ax.plot(
                    category_percentiles,  # Les percentiles sur l'axe X
                    category_positions,   # Les positions sur l'axe Y
                    marker="o", linestyle="-", color=category_colors[category],
                    label=category, zorder=4, linewidth=2
                )

        # Ajouter des titres par catégorie sur l'axe Y
        for category, color in category_colors.items():
            # Filtrer les tâches dans la catégorie
            category_data = data[data["Catégorie"] == category]
            
            # Si la catégorie n'est pas vide, ajouter un titre
            if not category_data.empty:
                # Calculer la position moyenne des tâches de la catégorie
                category_positions = category_data["Position"].tolist()
                mid_position = np.mean(category_positions)
                
                # Ajouter le texte pour le titre de la catégorie avec un cadre coloré
                ax.text(
                    x=-25,  # Décalage vers la gauche (en dehors des ticks Y)
                    y=mid_position,
                    s=category.upper(),
                    color="white",  # Couleur du texte
                    fontsize=20,
                    fontweight="bold",
                    ha="right",  # Aligner à droite
                    va="center", 
                    rotation=90,
                    bbox=dict(
                        facecolor=color,  # Couleur de fond
                        edgecolor=color,    # Couleur de la bordure (correspond à la catégorie)
                        boxstyle="round,pad=0.3",  # Bord arrondi avec padding
                        linewidth=2,         # Épaisseur de la bordure
                        alpha=1              # Transparence du fond
                    )
                )


        # Colorer les labels des ticks en fonction des catégories
        for idx, task_label in enumerate(ax.get_yticklabels()):
            if idx < len(data):
                task_category = data.iloc[idx]["Catégorie"]
                task_label.set_color(category_colors.get(task_category, "gray"))
                
        
        for spine in ["top", "right", "bottom", "left"]:
            ax.spines[spine].set_color("white")  # Couleur noire pour la bordure
            ax.spines[spine].set_linewidth(0)    # Épaisseur de la bordure

        # Supprimer la bordure noire à droite
        ax.spines["right"].set_visible(False)


        # Ajuster la mise en page
        plt.subplots_adjust(left=0.3, right=0.95, top=0.85, bottom=0.15)
        plt.tight_layout()

        # Afficher le graphique
        st.pyplot(fig)


# Ajouter la colonne "Catégorie" pour chaque tâche
def assign_category(task):
    for category, tasks in categories_mapping.items():
        if task in tasks:
            return category
    return "Autre"


# Étape 3 : Résultats
if st.session_state["scores_entered"]:
    st.header("Étape 3 : Résultats")

    age_data = st.session_state["age_data"]
    missing_norms = st.session_state["missing_norms"]


    if "Catégorie" not in age_data.columns:
        categories_mapping = {
            "Langage": [
                "Discrimination Phonologique", "Décision Lexicale Auditive",
                "Mots Outils", "Stock Lexical", "Compréhension Syntaxique", "Mots Outils - BOEHM"
            ],
            "Mémoire de Travail": [
                "Mémoire de travail verbale endroit empan", "Mémoire de travail verbale endroit brut",
                "Mémoire de travail verbale envers empan", "Mémoire de travail verbale envers brut",
                "Mémoire de travail non verbale endroit empan", "Mémoire de travail non verbale endroit brut",
                "Mémoire de travail non verbale envers empan", "Mémoire de travail non verbale envers brut"
            ],
            "Mise à jour": [
                "Mise à jour verbale empan", "Mise à jour verbale score",
                "Mise à jour non verbale empan", "Mise à jour non verbale score"
            ],
            "Inhibition": [
                "Inhibition verbale congruent score", "Inhibition verbale incongruent score",
                "Inhibition verbale congruent temps", "Inhibition verbale incongruent temps",
                "Inhibition verbale interférence score", "Inhibition verbale interférence temps",
                "Inhibition non verbale congruent score", "Inhibition non verbale incongruent score",
                "Inhibition non verbale congruent temps", "Inhibition non verbale incongruent temps",
                "Inhibition non verbale interférence score", "Inhibition non verbale interférence temps"
            ]
        }
        age_data["Catégorie"] = age_data["Tâche"].apply(assign_category)

    # Afficher le tableau des résultats
    st.write("")
    df_to_style = age_data.copy()
    
    # Supprimer la colonne "Catégorie"
    df_to_display = age_data.drop(columns=["Catégorie"]).reset_index(drop=True)
    def format_floats(value):
        if isinstance(value, float):
            return f"{value:.2f}".rstrip('0').rstrip('.')  # Arrondir à deux décimales et supprimer les zéros inutiles
        return value
    

    df_to_style = df_to_style.applymap(format_floats)  
    df_to_style["Percentile (%)"] = pd.to_numeric(age_data["Percentile (%)"], errors="coerce")

    #couleur sur les percentiles
    def color_percentiles_by_range(value):
        if pd.isna(value):  
            return ''  
        value = float(value)  
        if value <= 3:
            return 'background-color: rgba(212, 70, 70, 0.5); color: black;'  
        elif value <= 15:
            return 'background-color: rgba(245, 167, 47, 0.5); color: black;'  
        elif value <= 85:
            return 'background-color: rgba(96, 205, 114, 0.5); color: black;'  
        elif value <= 97:
            return 'background-color: rgba(141, 223, 155, 0.5); color: black;'  
        elif value <= 100:
            return 'background-color: rgba(174, 222, 182, 0.5); color: black;'  
        return ''  

    def color_task_text_by_category(row):
        category_colors = {
            "Langage": "#3798da",
            "Mémoire de Travail": "#eca113",
            "Mise à jour": "#e365d6",
            "Inhibition": "#8353da",
            "Autre": "gray"
        }
        category = row["Catégorie"]
        color = category_colors.get(category, "black")
        return [f"color: {color}; font-weight: bold;" if col == "Tâche" else "" for col in row.index]

    # Couleur percentiles
    styled_df = df_to_style.style.applymap(color_percentiles_by_range, subset=["Percentile (%)"])
    styled_df = styled_df.apply(color_task_text_by_category, axis=1)

    # Taille colonne
    col_config = {
        df_to_display.columns[0]: st.column_config.Column(width=300),  # Première colonne à 300
    }
    col_config.update({col: st.column_config.Column(width=100) for col in df_to_display.columns[1:]})  # Le reste à 100

    st.dataframe(styled_df, hide_index=True, use_container_width=True)


    # Sélection des tâches
    st.subheader("Sélectionnez les tâches à afficher dans le graphique")
    calculated_tasks = age_data[~age_data["Z-Score"].isna()]["Tâche"].tolist()
    tasks_by_category = {}
    for category, tasks in categories_mapping.items():
        tasks_in_category = [task for task in tasks if task in calculated_tasks]
        if tasks_in_category:
            tasks_by_category[category] = tasks_in_category

    col1, col2, col3, col4, col5 = st.columns([1, 2, 1, 2, 1])
    with col2:
        if st.button("Tout sélectionner"):
            selected_tasks = calculated_tasks
        else:
            selected_tasks = []

    with col4:
        if st.button("Tout désélectionner"):
            selected_tasks = []

    # Bouton sélection des tâches
    selected_tasks = st.multiselect(
        "Tâches calculées disponibles :", 
        options=calculated_tasks, 
        default=selected_tasks,
        help="Vous pouvez rechercher ou sélectionner des tâches dans la liste."
    )

# Sauvegarde graphique
    def save_styled_excel(dataframe):
        wb = Workbook()
        ws = wb.active
        ws.title = "Résultats"

        # remplissage percentiles
        fill_colors = {
            "red": "D44646",
            "orange": "F5A72F",
            "green": "60CD72",
            "light_green": "8DDF9B",
            "blue": "AEDFB6",
        }

        # couleurs catégories
        category_colors = {
            "Langage": "3798DA",
            "Mémoire de Travail": "ECA113",
            "Mise à jour": "E365D6",
            "Inhibition": "8353DA",
            "Autre": "808080",
        }

        #en-têtes
        headers = list(dataframe.columns)
        ws.append(headers)
        header_font = Font(bold=True)
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
            for cell in col:
                cell.font = header_font

        # Couleur excel
        for idx, row in dataframe.iterrows():
            ws.append(row.values.tolist())  
            excel_row = ws[idx + 2]  

            for col_idx, cell in enumerate(excel_row, start=1):
                
                # Couleur percentiles
                if headers[col_idx - 1] == "Percentile (%)":
                    try:
                        value = float(cell.value)
                        if value <= 3:
                            fill_color = fill_colors["red"]
                        elif value <= 15:
                            fill_color = fill_colors["orange"]
                        elif value <= 85:
                            fill_color = fill_colors["green"]
                        elif value <= 97:
                            fill_color = fill_colors["light_green"]
                        elif value <= 100:
                            fill_color = fill_colors["blue"]
                        else:
                            fill_color = None

                        if fill_color:
                            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    except (ValueError, TypeError):
                        pass  

                # couleur colone tâche
                if headers[col_idx - 1] == "Tâche":
                    category = dataframe.loc[idx, "Catégorie"]
                    color = category_colors.get(category, "000000")  
                    cell.font = Font(color=color, bold=True)
                
        buffer = io.BytesIO()
        try:
            wb.save(buffer)  
            buffer.seek(0)  
        except Exception as e:
            st.error(f"Erreur lors de la sauvegarde du fichier Excel : {e}")
            return None  

        return buffer


    def save_graph_and_data(data, selected_tasks):
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w") as zf:
            
            # graphique
            fig_buffer = io.BytesIO()
            plot_grouped_scores(data, selected_tasks)  
            plt.savefig(fig_buffer, format='png', dpi=300, bbox_inches="tight")
            plt.close()  
            fig_buffer.seek(0)
            zf.writestr(f"{st.session_state['child_id']}_Graphique_Comprendre.png", fig_buffer.read())

            # Excel
            styled_excel = save_styled_excel(data) 
            if styled_excel: 
                zf.writestr(f"{st.session_state['child_id']}_Tableau_Comprendre.xlsx", styled_excel.read())
            else:
                st.warning("Le fichier Excel n'a pas pu être généré et ne sera pas inclus dans l'archive.")

        buffer.seek(0)
        return buffer


if st.session_state["scores_entered"] and selected_tasks:
        st.subheader("Téléchargez les résultats")
        zip_file = save_graph_and_data(age_data, selected_tasks)
        st.download_button(
            label="📥 Télécharger le tableau des résultats et le graphique (ZIP)",
            data=zip_file,
            file_name=f"{st.session_state['child_id']}_Resultats_Comprendre.zip",
            mime="application/zip"
        )
